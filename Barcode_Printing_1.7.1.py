import qrcode
import os
import xlsxwriter
import datetime
import openpyxl
from openpyxl import load_workbook
from PIL import Image, ImageFont, ImageDraw ,ImageWin
import win32print
import win32ui

name_of_workbook= input("name for the new excel sheet : ")                       #A new excel sheet in made and the name for the new sheet is asked from the user 
name_of_workbook_1=name_of_workbook+".xlsx"

new_workbook = xlsxwriter.Workbook(name_of_workbook_1)
new_sheet = new_workbook.add_worksheet()

new_sheet.write("A1","S.no")
new_sheet.write("B1","Date")
new_sheet.write("C1","Time")
new_sheet.write("D1","Serial No.")
new_sheet.write("E1","Comments")
new_workbook.close()

wb = openpyxl.load_workbook(name_of_workbook_1)
ws = wb.active

i=2
j=1
while True:
    
    img= input("please scan the QRcode/Barcode :")				#the user is asked to enter the data whose Qr code is to be genrated, the entry can be manual or using a barcode scanner.
    date=datetime.datetime.now()
    date_1=date.strftime("%d-%m-%y")
    time=date.strftime("%H:%M:%S")
    image=qrcode.QRCode(version=1,						#the style of the Qrcode can be edited here.
                        error_correction=qrcode.constants.ERROR_CORRECT_L,
                        box_size=5,
                        border=5)
    image.add_data(img)
    image.make(fit="False")
    img_1= image.make_image(fill_color="black", back_color="white")
    img_1.save("Miko QrCode.jpg")						#the name with which the Qrcode is saved as image.
    ws.cell(row=i,column=1,value=j)
    ws.cell(row=i,column=2,value=date_1)
    ws.cell(row=i,column=3,value=time)
    ws.cell(row=i,column=4,value=img)
    i=i+1
    j=j+1
    wb.save(name_of_workbook_1)
    my_image = Image.open("Miko QrCode.jpg")
    title_font = ImageFont.truetype("ariblk.ttf", 13)			#the font style and size is set here.
    image_editable = ImageDraw.Draw(my_image)
    image_editable.text((0,155), img,"black", font=title_font)		#the position and color of the test is set here.
    my_image.save("Miko QrCode_1.jpg")
    print("QRcode Generated")
    
    print("Printing Barcode")
    HORZRES = 8								#From here it is the code for automated printing so no dialogue box for page set-uyp and printer selection comes.
    VERTRES = 10
    LOGPIXELSX = 88
    LOGPIXELSY = 90
    PHYSICALWIDTH = 110
    PHYSICALHEIGHT = 111
    PHYSICALOFFSETX = 112
    PHYSICALOFFSETY = 113

    printer_name = win32print.GetDefaultPrinter ()
    file_name = "Miko QrCode_1.jpg"
    hDC = win32ui.CreateDC ()
    hDC.CreatePrinterDC (printer_name)
    printable_area = hDC.GetDeviceCaps (HORZRES), hDC.GetDeviceCaps (VERTRES)
    printer_size = hDC.GetDeviceCaps (PHYSICALWIDTH), hDC.GetDeviceCaps (PHYSICALHEIGHT)
    printer_margins = hDC.GetDeviceCaps (PHYSICALOFFSETX), hDC.GetDeviceCaps (PHYSICALOFFSETY)
    bmp = Image.open (file_name)
    if bmp.size[0] > bmp.size[1]:
      bmp = bmp.rotate (90)

    ratios = [1.0 * printable_area[0] / bmp.size[0], 1.0 * printable_area[1] / bmp.size[1]]
    scale = min (ratios)
    hDC.StartDoc (file_name)
    hDC.StartPage ()

    dib = ImageWin.Dib (bmp)
    scaled_width, scaled_height = [int (scale * i) for i in bmp.size]
    x1 = int ((printer_size[0] - scaled_width) / 2)
    y1 = int ((printer_size[1] - scaled_height) / 2)
    x2 = x1 + scaled_width
    y2 = y1 + scaled_height
    dib.draw (hDC.GetHandleOutput (), (x1, y1, x2, y2))

    hDC.EndPage ()
    hDC.EndDoc ()
    hDC.DeleteDC ()


