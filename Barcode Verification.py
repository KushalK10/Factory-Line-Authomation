import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

bookName = r'Name_of _excel.xlsx'
sheetName = 'Sheet1'

Font1= Font(name="Arial",size=12, color="FF6347",bold=True)
Font2= Font(name="Arial",size=12, color="07F556",bold=True)

wb = load_workbook(bookName)
ws = wb.active
while True :
    search=input('Please scan the Barcode : ')			#the input that needs to be check in the invetory, input can come from a barcode scanner or manually from a keyboard
    for row in ws.iter_rows(2,50502):					#in the bracket add the first and the last row to be in the iterartion loop.	
        for cell in row:
            if cell.value == search:
                if cell.font == Font2:
                    cell.font = Font1
                    print('Repeated Barcode')
                    wb.save("Name_of _excel.xlsx")			#type the name of the excel sheet which is being cross checked with
                    break
                elif cell.font == Font1:
                    print('Repeated Barcode')
                    wb.save("Name_of _excel.xlsx") 			#type the name of the excel sheet which is being cross checked with
                    break                    
                else:
                    cell.font = Font2
                    print('Verified, go ahead!')
                    wb.save("Name_of _excel.xlsx")			#type the name of the excel sheet which is being cross checked with
                    break
                
                    
